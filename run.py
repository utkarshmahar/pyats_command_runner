from webexteamssdk import WebexTeamsAPI
import os
import sys
import argparse
import yaml
import openpyxl,requests
from datetime import datetime
now = datetime.now()

dt_string = now.strftime("%d-%m-%H-%M-%S")


parser = argparse.ArgumentParser()
#parser.add_argument('-t', '--testbed', type=str, required=False, help="Testbed")
#parser.add_argument('-d', '--data', type=str, required=False, help="Data")
parser.add_argument('-wb', '--workbook', type=str, required=True, help="Workbook Name")
parser.add_argument('-ws', '--worksheet', type=str, required=True, help="Worksheet Name")
parser.add_argument('-w', '--webex', type=str, required=False, help="Webex")
parser.add_argument('-ssh','--ssh_options', type=str, required=False, help="SSH Options for old ciphers")
args = parser.parse_args()
''' 
if args.testbed != None:
 testbed = "testbeds/"+args.testbed
if args.data != None:
 data_file = "data_files/"+args.data
'''
input_file = args.workbook
sheet = args.worksheet
ssh_options = args.ssh_options
file_name = input_file.split(".xlsx")[0]
job_name = f"{file_name}_{sheet}__{dt_string}"


def process_testbed(input_file_path):
    global testbed 
    output_file_path = input_file_path.split(".xlsx")[0]+"_testbed.xlsx"
    testbed = output_file_path
    try:
        # Load the input workbook
        workbook = openpyxl.load_workbook(input_file_path)

        # Specify the worksheet name
        input_sheet_name = "testbed"

        # Get the input sheet
        sheet = workbook[input_sheet_name]

        # Validate headers
        expected_headers = ['hostname', 'ip', 'os']
        actual_headers = [cell.value.lower() for cell in sheet[1]]
        
        if expected_headers != actual_headers:
            raise ValueError("Invalid headers in the input file")

        # Create a new workbook and sheet
        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.active

        # Write new headers
        new_headers = ['hostname', 'ip', 'username', 'password', 'protocol', 'os']
        for col_num, header in enumerate(new_headers, start=1):
            new_sheet.cell(row=1, column=col_num, value=header)

        # Copy data with default values for username, password, protocol, and copy os from input
        for row_num in range(2, sheet.max_row + 1):
            new_sheet.cell(row=row_num, column=1, value=sheet.cell(row=row_num, column=1).value)  # Hostname
            new_sheet.cell(row=row_num, column=2, value=sheet.cell(row=row_num, column=2).value)  # IP
            new_sheet.cell(row=row_num, column=3, value='do not leave blank')  # Username
            new_sheet.cell(row=row_num, column=4, value='do not leave blank')  # Password
            new_sheet.cell(row=row_num, column=5, value='ssh')  # Protocol
            new_sheet.cell(row=row_num, column=6, value=sheet.cell(row=row_num, column=3).value)  # OS

        # Save the new workbook
        new_workbook.save("testbeds/"+output_file_path)
        print("Testbed Excel created")
      

    except Exception as e:
        print(f"Error: {e}")


def process_data_file(excel_file_path, sheet_name):
    global data_file
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(excel_file_path)

        # Get the specified worksheet
        sheet = workbook[sheet_name]

        # Validate A1 cell
        if sheet['A1'].value != 'Devices':
            raise ValueError("Invalid A1 cell value. It should be 'Devices'.")

        # Read values in row A starting from A2 until a blank cell
        devices_array = []
        row_num = 2
        while sheet.cell(row=row_num, column=1).value is not None:
            devices_array.append(sheet.cell(row=row_num, column=1).value)
            row_num += 1

        # Find the index of the first blank row after "Devices"
        blank_row_index = None
        for row_num in range(row_num, sheet.max_row + 1):
            if sheet.cell(row=row_num, column=1).value is None:
                blank_row_index = row_num
                break

        if blank_row_index is None:
            raise ValueError("Blank row not found after 'Devices'.")

        # Check if the next row contains "Commands" in the first column
        if sheet.cell(row=blank_row_index + 1, column=1).value != 'Commands':
            raise ValueError("Row after 'Devices' does not contain 'Commands' in the first column.")

        # Read values in the "Commands" column and subsequent columns until a blank cell
        commands_array = []
        row_num = blank_row_index + 1
        while sheet.cell(row=row_num, column=1).value is not None:
            if sheet.cell(row=row_num, column=1).value != 'Commands':
                commands_array.append(sheet.cell(row=row_num, column=1).value)
            row_num += 1

        # Create a dictionary for YAML
        yaml_data = {
            'parameters': {
                'network_devices': devices_array,
                'commands': commands_array
            }
        }

        # Save the dictionary as YAML
        yaml_file_path = f"data_files/{sheet_name}.yaml"
        with open(yaml_file_path, 'w') as yaml_file:
            yaml.dump(yaml_data, yaml_file, default_flow_style=False)
    
        print(f"YAML file '{yaml_file_path}' created successfully!")
        data_file = yaml_file_path
        
    except Exception as e:
        print(f"Error: {e}")
def build_testbed() :
    global testbed
    
    if ".xlsx" in testbed or ".csv" in testbed:
        new_testbed = "testbeds/" + testbed.split(".")[0] + ".yaml"
        os.system(f"pyats create testbed file --path testbeds/{testbed} --output {new_testbed}")
        with open(new_testbed) as r:
            dict = yaml.safe_load(r)
            dict["testbed"] = {'credentials': {'default': {'username': "%ASK{}", 'password': "%ASK{}"}}}
        for device, device_data in dict["devices"].items():
            device_data.pop('credentials')
            if ssh_options :
             device_data["connections"]["cli"]["ssh_options"] = ssh_options
        linux_device = {'local_linux': {'os': 'linux', 'type': 'linux', 'connections': {
            'cli': {'protocols': 'ssh', 'ip': '127.0.0.1', 'command': 'bash'}}}}
        dict['devices'].update(linux_device)
        with open(new_testbed, 'w') as w:
            yaml.dump(dict, w)
        testbed = new_testbed
        
def build_data_file():
    with open(data_file,'r') as d:
        contents = d.read()
    with open("data_files/temp_data.yaml",'a') as td:
        td.write("extends: scripts/data.yaml \n")
        td.write(contents)

def cleanup():
    os.system("cp " + "results/"+job_name+"/TaskLog.job.html " + "results/" +job_name+".html")
    os.system("rm -r "+"results/"+job_name)
    os.system("rm data_files/temp_data.yaml")
def publish_webex():
     if args.webex :
         with open(args.webex, 'r') as f:
             o = yaml.full_load(f)["WEBEX"]
         try:
             webex_api = WebexTeamsAPI(access_token=o["WEBEX_BOT_TOKEN"], proxies={"https": o["PROXY"]})
         except :
             webex_api = WebexTeamsAPI(access_token=o["WEBEX_BOT_TOKEN"])
       
         webex_api.messages.create(roomId=o["WEBEX_ROOM"], markdown="Publishing Test Result " )
         webex_api.messages.create(roomId=o["WEBEX_ROOM"], files=['results/'+job_name+".html"])

def download_file():
     try:

         with open(args.webex, 'r') as f:    
          o = yaml.full_load(f)["WEBEX"]
          try: 
           api = WebexTeamsAPI(access_token=o["WEBEX_BOT_TOKEN"])
          except:
           api = WebexTeamsAPI(access_token=o["WEBEX_BOT_TOKEN"])  
          messages = api.messages.list(roomId=o["1_1_BOT_ROOM_ID"])
          for message in messages:
            if message.files:
              print(message)
              break
          headers = { 'Authorization': f'Bearer {o["WEBEX_BOT_TOKEN"]}'} 
          file_url =  message.files[0]
          response = requests.get(file_url, headers=headers, stream=True)
          file_name = message.text.split(" ")[1]
          with open(file_name,"wb") as file:
           file.write(response.content)
          print(f"Downloaded file {file_name}")
     except Exception as e :
           print(e)
           print(f"Couldn't download file from webex room")

if __name__ == '__main__':
    if args.webex :
        download_file()
    process_testbed(input_file)
    process_data_file(input_file,sheet)
    build_testbed()
    build_data_file()
    os.system(f"pyats run job scripts/job.py --testbed {testbed} --datafile {data_file} --html-logs results/{job_name}")
    cleanup()
    publish_webex()